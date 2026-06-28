from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font


TEMPLATE_PATH = Path("template.xlsx")
ITEM_START_ROW = 10
TAMIL_FONT = Font(name="JF Kamban", size=11)


def generate_bill_excel(bill_data: dict[str, Any]) -> BytesIO:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook.active

    sheet["B4"] = bill_data["customer_name"]
    sheet["F4"] = bill_data["bill_id"]
    sheet["F5"] = datetime.now().strftime("%d-%m-%Y")

    for index, item in enumerate(bill_data["items"], start=ITEM_START_ROW):
        product_name = item["product_name"]
        quantity = item["quantity"]
        price = item["price"]

        product_cell = sheet.cell(row=index, column=1)
        product_cell.value = product_name
        product_cell.font = TAMIL_FONT

        sheet.cell(row=index, column=4).value = quantity
        sheet.cell(row=index, column=5).value = price
        sheet.cell(row=index, column=6).value = quantity * price

    sheet["E30"] = "Packaging Fee"
    sheet["F30"] = bill_data["packaging_fee"]
    sheet["E32"] = "Grand Total"
    sheet["F32"] = bill_data["grand_total"]
    sheet["E34"] = "Advance Paid"
    sheet["F34"] = bill_data["advance"]
    sheet["E36"] = "Balance Due"
    sheet["F36"] = bill_data["balance"]

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
