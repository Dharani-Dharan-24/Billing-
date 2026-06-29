from openpyxl import Workbook
from datetime import datetime, time, timedelta
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any
from typing import Literal

from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from pydantic import BaseModel, Field
from sqlalchemy import func, inspect, select, text
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session, selectinload

from database import Base, engine, get_db
from models import Bill, BillItem, Customer, Product

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="Billing API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory="static"), name="static")
INDEX_HTML = Path("templates/index.html")
TEMPLATE_PATH = Path("template.xlsx")

pdfmetrics.registerFont(TTFont('JFKamban', 'static/fonts/jf-kamban.ttf'))


class ItemSchema(BaseModel):
    """Strict validation schema for bill line items with all required fields."""
    product_name: str = Field(min_length=1)
    quantity: float = Field(gt=0, description="Quantity must be greater than 0")
    price: float = Field(ge=0, description="Price must be non-negative")
    total: float = Field(ge=0, description="Line total (quantity × price)")

    class Config:
        validate_assignment = True


class ItemInput(BaseModel):
    product_name: str
    quantity: float = Field(gt=0)
    price: float = Field(ge=0)


class BillPayload(BaseModel):
    """Strict validation schema for complete bill payload."""
    customer_name: str = Field(min_length=1)
    customer_type: Literal["New", "Regular"]
    location: str | None = None
    packaging_fee: float = Field(ge=0)
    advance_paid: float = Field(ge=0)
    items: list[ItemSchema] = Field(min_length=1)

    class Config:
        validate_assignment = True


class BillRequest(BaseModel):
    customer_name: str
    customer_type: Literal["New", "Regular"]
    phone: str | None = None
    location: str | None = None
    items: list[ItemInput] = Field(min_length=1)
    packaging_fee: float = Field(ge=0)
    advance_paid: float = Field(ge=0)


class CustomerResponse(BaseModel):
    id: int
    name: str
    customer_type: str
    phone: str | None
    location: str | None


class CustomerCreate(BaseModel):
    name: str
    phone: str | None = None
    location: str | None = None
    customer_type: str


class CustomerUpdate(BaseModel):
    name: str | None = None
    phone: str | None = None
    location: str | None = None
    customer_type: str | None = None


class ProductCreate(BaseModel):
    name: str
    default_price: float = Field(ge=0)


class ProductUpdate(BaseModel):
    name: str | None = None
    default_price: float | None = Field(default=None, ge=0)


class ProductResponse(BaseModel):
    id: int
    name: str
    default_price: float


class BillHistoryResponse(BaseModel):
    bill_id: str
    customer_name: str
    grand_total: float
    created_at: datetime


@app.on_event("startup")
def create_tables() -> None:
    Base.metadata.create_all(bind=engine)
    apply_schema_updates()


@app.get("/", response_class=HTMLResponse)
def index() -> HTMLResponse:
    return HTMLResponse(INDEX_HTML.read_text(encoding="utf-8"))


def customer_to_response(customer: Customer) -> dict[str, object]:
    return {
        "id": customer.id,
        "name": customer.name,
        "customer_type": customer.customer_type,
        "phone": customer.phone,
        "location": customer.location,
    }


def product_to_response(product: Product) -> dict[str, object]:
    return {
        "id": product.id,
        "name": product.name,
        "default_price": product.default_price,
    }


def bill_to_history_response(bill: Bill) -> dict[str, object]:
    customer_name = bill.customer.name if bill.customer is not None else ""
    return {
        "bill_id": bill.bill_id,
        "customer_name": customer_name,
        "grand_total": bill.grand_total,
        "created_at": bill.created_at,
    }


def apply_schema_updates() -> None:
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    with engine.begin() as connection:
        if "customers" in table_names:
            customer_columns = {column["name"] for column in inspector.get_columns("customers")}
            if "location" not in customer_columns:
                connection.execute(text("ALTER TABLE customers ADD COLUMN location VARCHAR"))

        if "bill_items" not in table_names:
            bill_items_table: Any = BillItem.__table__
            bill_items_table.create(bind=connection, checkfirst=True)


def format_customer_name(name: str) -> str:
    normalized = name.strip()
    lowered = normalized.lower()
    if lowered.startswith("mr. ") or lowered.startswith("ms. "):
        return normalized
    return f"Mr./Ms. {normalized}"


def generate_bill_excel(bill_data: dict[str, Any]) -> BytesIO:
    # 1) Create a fresh workbook
    wb = Workbook()
    ws = wb.active

    # 2) Column dimensions
    ws.column_dimensions["A"].width = 8   # Sl.No
    ws.column_dimensions["B"].width = 45  # ITEM NAME
    ws.column_dimensions["C"].width = 10  # QTY
    ws.column_dimensions["D"].width = 15  # Rate/unit
    ws.column_dimensions["E"].width = 15  # Total

    # 3) Reusable style objects
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    standard_font = Font(name='Times New Roman', size=16)
    bold_font = Font(name='Times New Roman', size=16, bold=True)
    tamil_font = Font(name='JF Kamban', size=16)
    money_format = '#,##0.00'

    # ---------------------------------------------------------------------
    # 4) Static Header Section (Rows 1‑21) – leave rows 1‑13 empty for spacing
    # ---------------------------------------------------------------------
    # Row 14, Column E – Date
    date_cell = ws.cell(row=14, column=5)
    date_val = bill_data.get('bill_date') or datetime.now()
    date_cell.value = date_val
    date_cell.number_format = 'DD.MM.YYYY'
    date_cell.font = bold_font
    date_cell.alignment = right_align

    # Row 18, Column A – "To,"
    ws.cell(row=18, column=1, value='To,').font = standard_font

    # Row 19, Column B – Customer name (bold)
    ws.cell(row=19, column=2, value=str(bill_data.get('customer_name', ''))).font = bold_font

    # Row 20, Column B – Location or Phone (whichever is present)
    loc = bill_data.get('location')
    ph = bill_data.get('phone')
    ws.cell(row=20, column=2, value=loc or ph or '').font = standard_font

    # ---------------------------------------------------------------------
    # 5) Grid Header (Row 22)
    # ---------------------------------------------------------------------
    headers = ['Sl.No', 'ITEM NAME', 'QTY', 'Rate/unit', 'Total']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=22, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = left_align if col_idx == 2 else center_align

    # ---------------------------------------------------------------------
    # 6) Dynamic Item Grid (starts Row 23)
    # ---------------------------------------------------------------------
    items = bill_data.get('items', [])
    start_row = 23
    for idx, item in enumerate(items):
        r = start_row + idx
        # Column 1 – Index
        c1 = ws.cell(row=r, column=1, value=idx + 1)
        c1.font = standard_font
        c1.alignment = center_align
        c1.border = thin_border
        # Column 2 – Item name (Tamil font, left)
        c2 = ws.cell(row=r, column=2, value=str(item.get('product_name', '')))
        c2.font = tamil_font
        c2.alignment = left_align
        c2.border = thin_border
        # Column 3 – Quantity (center)
        c3 = ws.cell(row=r, column=3, value=float(item.get('quantity', 1)))
        c3.font = standard_font
        c3.alignment = center_align
        c3.border = thin_border
        # Column 4 – Rate/unit (right, money format)
        c4 = ws.cell(row=r, column=4, value=float(item.get('price', 0.0)))
        c4.font = standard_font
        c4.alignment = right_align
        c4.border = thin_border
        c4.number_format = money_format
        # Column 5 – Total (right, money format)
        total_val = float(item.get('total', float(item.get('quantity', 1)) * float(item.get('price', 0.0))))
        c5 = ws.cell(row=r, column=5, value=total_val)
        c5.font = standard_font
        c5.alignment = right_align
        c5.border = thin_border
        c5.number_format = money_format

    # ---------------------------------------------------------------------
    # 7) Summary Block (dynamic positioning)
    # ---------------------------------------------------------------------
    summary_start = start_row + len(items) + 2  # two empty rows after the grid
    summary_data = [
        ("Packing Charges", float(bill_data.get('packaging_fee', 0.0)), False),
        ("TOTAL",            float(bill_data.get('grand_total', 0.0)), True ),
        ("Advance",          float(bill_data.get('advance_paid', 0.0)), False),
        ("Balance Amount",   float(bill_data.get('balance_due', 0.0)), True ),
    ]
    for offset, (label_text, amount, is_bold) in enumerate(summary_data):
        row = summary_start + offset * 2  # blank row between each entry
        # Merge C and D columns for the label
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        
        # Column C – label (merged with D)
        label_cell = ws.cell(row=row, column=3, value=label_text)
        label_cell.font = bold_font if is_bold else standard_font
        label_cell.alignment = left_align
        label_cell.border = thin_border
        
        # Column D - need border to complete the merged cell outline
        ws.cell(row=row, column=4).border = thin_border
        
        # Column E – amount
        amount_cell = ws.cell(row=row, column=5, value=amount)
        amount_cell.font = bold_font if is_bold else standard_font
        amount_cell.alignment = right_align
        amount_cell.border = thin_border
        amount_cell.number_format = money_format

    # ---------------------------------------------------------------------
    # 8) Return the workbook as a stream
    # ---------------------------------------------------------------------
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@app.post(
    "/api/customers",
    response_model=CustomerResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_customer(
    request: CustomerCreate,
    db: Session = Depends(get_db),
) -> dict[str, object]:
    customer = Customer(
        name=request.name,
        phone=request.phone,
        location=request.location,
        customer_type=request.customer_type,
    )
    db.add(customer)

    try:
        db.commit()
        db.refresh(customer)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Customer name already exists.",
        ) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not create customer.",
        ) from exc

    return customer_to_response(customer)


@app.get("/api/customers", response_model=list[CustomerResponse])
def get_customers(db: Session = Depends(get_db)) -> list[dict[str, object]]:
    customers = db.scalars(select(Customer).order_by(Customer.name)).all()
    return [customer_to_response(customer) for customer in customers]


@app.put("/api/customers/{customer_id}", response_model=CustomerResponse)
def update_customer(
    customer_id: int,
    request: CustomerUpdate,
    db: Session = Depends(get_db),
) -> dict[str, object]:
    customer = db.get(Customer, customer_id)
    if customer is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Customer not found.",
        )

    if request.name is not None:
        customer.name = request.name
    if request.phone is not None:
        customer.phone = request.phone
    if request.location is not None:
        customer.location = request.location
    if request.customer_type is not None:
        customer.customer_type = request.customer_type

    try:
        db.commit()
        db.refresh(customer)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Customer name already exists.",
        ) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not update customer.",
        ) from exc

    return customer_to_response(customer)


@app.delete("/api/customers/{customer_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_customer(customer_id: int, db: Session = Depends(get_db)) -> Response:
    customer = db.get(Customer, customer_id)
    if customer is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Customer not found.",
        )

    db.delete(customer)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Customer cannot be deleted because bills are linked to it.",
        ) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not delete customer.",
        ) from exc

    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.post(
    "/api/products",
    response_model=ProductResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_product(
    request: ProductCreate,
    db: Session = Depends(get_db),
) -> dict[str, object]:
    product = Product(name=request.name, default_price=request.default_price)
    db.add(product)

    try:
        db.commit()
        db.refresh(product)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Product name already exists.",
        ) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not create product.",
        ) from exc

    return product_to_response(product)


@app.get("/api/products", response_model=list[ProductResponse])
def get_products(db: Session = Depends(get_db)) -> list[dict[str, object]]:
    products = db.scalars(select(Product).order_by(Product.name)).all()
    return [product_to_response(product) for product in products]


@app.get("/api/bills", response_model=list[BillHistoryResponse])
def get_bills(db: Session = Depends(get_db)) -> list[dict[str, object]]:
    bills = db.scalars(
        select(Bill)
        .options(selectinload(Bill.customer))
        .order_by(Bill.created_at.desc())
    ).all()
    return [bill_to_history_response(bill) for bill in bills]


@app.put("/api/products/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: int,
    request: ProductUpdate,
    db: Session = Depends(get_db),
) -> dict[str, object]:
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found.",
        )

    if request.name is not None:
        product.name = request.name
    if request.default_price is not None:
        product.default_price = request.default_price

    try:
        db.commit()
        db.refresh(product)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Product name already exists.",
        ) from exc
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not update product.",
        ) from exc

    return product_to_response(product)


@app.delete("/api/products/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(product_id: int, db: Session = Depends(get_db)) -> Response:
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found.",
        )

    db.delete(product)
    try:
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Could not delete product.",
        ) from exc

    return Response(status_code=status.HTTP_204_NO_CONTENT)


def generate_pdf_bill(payload: dict) -> BytesIO:
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4, topMargin=170, leftMargin=45, rightMargin=45, bottomMargin=50)
    elements = []
    
    data = []
    date_str = f"{payload.get('bill_date', datetime.now()).strftime('%d-%m-%Y')}"
    data.append(["", "", "", "", date_str])
    data.append(["", "", "", "", ""])
    data.append(["", "", "", "", ""])
    data.append([f"To, {payload.get('customer_name', '')}", "", "", "", ""])
    data.append([f"{payload.get('location', '') or ''}", "", "", "", ""])
    data.append(["", "", "", "", ""])
    data.append(["Sl.No", "ITEM NAME", "QTY", "Rate/unit", "Total"])
    
    for idx, item in enumerate(payload.get('items', [])):
        data.append([
            str(idx + 1),
            item.get('product_name', ''),
            str(item.get('quantity', 0)),
            f"{item.get('price', 0):.2f}",
            f"{item.get('total', 0):.2f}"
        ])
    
    data.append(["", "Packing Charges", "", "", f"{payload.get('packaging_fee', 0):.2f}"])
    data.append(["", "TOTAL", "", "", f"{payload.get('grand_total', 0):.2f}"])
    data.append(["", "Advance", "", "", f"{payload.get('advance_paid', 0):.2f}"])
    data.append(["", "Balance Amount", "", "", f"{payload.get('balance_due', 0):.2f}"])
    
    header_idx = 6
    last_item_idx = header_idx + len(payload.get('items', []))
    
    style = TableStyle([
        ('ALIGN', (0, header_idx), (-1, -1), 'LEFT'),
        ('GRID', (0, header_idx), (-1, last_item_idx), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (1, header_idx + 1), (1, last_item_idx), 'JFKamban'),
        ('ALIGN', (0, header_idx), (0, -1), 'CENTER'),
        ('ALIGN', (2, header_idx), (2, -1), 'CENTER'),
        ('ALIGN', (3, header_idx), (3, -1), 'RIGHT'),
        ('ALIGN', (4, header_idx), (4, -1), 'RIGHT'),
        ('ALIGN', (4, 0), (4, 0), 'RIGHT'),
        
        # Neat spacing padding for the table
        ('TOPPADDING', (0, header_idx), (-1, -1), 6),
        ('BOTTOMPADDING', (0, header_idx), (-1, -1), 6),

        # Spanning and right-aligning the summary labels
        ('SPAN', (1, last_item_idx + 1), (3, last_item_idx + 1)),
        ('SPAN', (1, last_item_idx + 2), (3, last_item_idx + 2)),
        ('SPAN', (1, last_item_idx + 3), (3, last_item_idx + 3)),
        ('SPAN', (1, last_item_idx + 4), (3, last_item_idx + 4)),
        ('ALIGN', (1, last_item_idx + 1), (3, last_item_idx + 4), 'LEFT'),
        ('LEFTPADDING', (1, last_item_idx + 1), (3, last_item_idx + 4), 270),

        # Bold 'TOTAL' and 'Balance Amount'
        ('FONTNAME', (1, last_item_idx + 2), (1, last_item_idx + 2), 'Helvetica-Bold'),
        ('FONTNAME', (4, last_item_idx + 2), (4, last_item_idx + 2), 'Helvetica-Bold'),
        ('FONTNAME', (1, last_item_idx + 4), (1, last_item_idx + 4), 'Helvetica-Bold'),
        ('FONTNAME', (4, last_item_idx + 4), (4, last_item_idx + 4), 'Helvetica-Bold'),
    ])
    
    t = Table(data, colWidths=[40, 200, 60, 100, 100])
    t.setStyle(style)
    elements.append(t)
    
    doc.build(elements)
    stream.seek(0)
    return stream


@app.post("/api/generate-pdf-bill")
def generate_pdf_bill_endpoint(payload: BillPayload, db: Session = Depends(get_db)) -> StreamingResponse:
    items_with_validation = []
    for idx, item in enumerate(payload.items):
        calculated_total = item.quantity * item.price
        if abs(calculated_total - item.total) > 0.01:
            raise HTTPException(status_code=400, detail=f"Item {idx + 1}: Total mismatch.")
        items_with_validation.append(ItemInput(product_name=item.product_name, quantity=item.quantity, price=item.price))

    request = BillRequest(
        customer_name=payload.customer_name,
        customer_type=payload.customer_type,
        phone=None,
        location=payload.location,
        items=items_with_validation,
        packaging_fee=payload.packaging_fee,
        advance_paid=payload.advance_paid,
    )

    today = datetime.now().date()
    start_of_day = datetime.combine(today, time.min)
    start_of_tomorrow = start_of_day + timedelta(days=1)
    bills_today = db.scalar(select(func.count()).select_from(Bill).where(Bill.created_at >= start_of_day).where(Bill.created_at < start_of_tomorrow))
    sequence_number = (bills_today or 0) + 1
    bill_id = f"BILL-{today.strftime('%d%m%Y')}-{sequence_number:03d}"

    subtotal = sum(item.quantity * item.price for item in request.items)
    grand_total = subtotal + request.packaging_fee
    balance_due = grand_total - request.advance_paid

    customer = db.scalar(select(Customer).where(Customer.name == request.customer_name))
    if request.customer_type == "New":
        if customer is not None:
            raise HTTPException(status_code=409, detail="Customer already exists.")
        customer = Customer(name=request.customer_name, customer_type=request.customer_type, phone=None, location=request.location)
        db.add(customer)
        db.flush()
    elif customer is None:
        raise HTTPException(status_code=404, detail="Regular customer not found.")
    else:
        if request.location is not None:
            customer.location = request.location

    bill = Bill(
        bill_id=bill_id, customer_id=customer.id, subtotal=subtotal, packaging_fee=request.packaging_fee,
        grand_total=grand_total, advance_paid=request.advance_paid, balance_due=balance_due, created_at=datetime.now()
    )
    db.add(bill)
    db.add_all([BillItem(bill_id=bill_id, product_name=str(item.product_name).strip(), quantity=float(item.quantity), price=float(item.price), total=float(item.quantity * item.price)) for item in request.items])
    db.commit()

    stream = generate_pdf_bill({
        "bill_id": bill_id,
        "customer_name": customer.name,
        "customer_type": customer.customer_type,
        "phone": customer.phone,
        "location": customer.location,
        "bill_date": datetime.now(),
        "items": [{"product_name": item.product_name, "quantity": item.quantity, "price": item.price, "total": item.total} for item in payload.items],
        "subtotal": subtotal,
        "packaging_fee": request.packaging_fee,
        "grand_total": grand_total,
        "advance_paid": request.advance_paid,
        "balance_due": balance_due,
    })

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{bill_id}.pdf"'},
    )


@app.post("/api/generate-bill")
def generate_bill(payload: BillPayload, db: Session = Depends(get_db)) -> StreamingResponse:
    """Generate bill with strict Pydantic validation of incoming payload."""
    # Convert BillPayload to BillRequest format for processing
    # Calculate totals and validate items
    items_with_validation = []
    for idx, item in enumerate(payload.items):
        # Verify total calculation matches frontend
        calculated_total = item.quantity * item.price
        if abs(calculated_total - item.total) > 0.01:  # Allow small floating-point differences
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Item {idx + 1}: Total mismatch. Expected {calculated_total}, got {item.total}",
            )
        items_with_validation.append(
            ItemInput(
                product_name=item.product_name,
                quantity=item.quantity,
                price=item.price,
            )
        )

    # Convert to BillRequest for existing logic
    request = BillRequest(
        customer_name=payload.customer_name,
        customer_type=payload.customer_type,
        phone=None,
        location=payload.location,
        items=items_with_validation,
        packaging_fee=payload.packaging_fee,
        advance_paid=payload.advance_paid,
    )

    today = datetime.now().date()
    start_of_day = datetime.combine(today, time.min)
    start_of_tomorrow = start_of_day + timedelta(days=1)

    bills_today = db.scalar(
        select(func.count())
        .select_from(Bill)
        .where(Bill.created_at >= start_of_day)
        .where(Bill.created_at < start_of_tomorrow)
    )
    sequence_number = (bills_today or 0) + 1
    bill_id = f"BILL-{today.strftime('%d%m%Y')}-{sequence_number:03d}"

    subtotal = sum(item.quantity * item.price for item in request.items)
    grand_total = subtotal + request.packaging_fee
    balance_due = grand_total - request.advance_paid

    customer = db.scalar(
        select(Customer).where(Customer.name == request.customer_name)
    )

    if request.customer_type == "New":
        if customer is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Customer already exists.",
            )
        customer = Customer(
            name=request.customer_name,
            customer_type=request.customer_type,
            phone=None,
            location=request.location,
        )
        db.add(customer)
        db.flush()
    elif customer is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Regular customer not found.",
        )
    else:
        if request.location is not None:
            customer.location = request.location

    bill = Bill(
        bill_id=bill_id,
        customer_id=customer.id,
        subtotal=subtotal,
        packaging_fee=request.packaging_fee,
        grand_total=grand_total,
        advance_paid=request.advance_paid,
        balance_due=balance_due,
        created_at=datetime.now(),
    )
    db.add(bill)

    db.add_all(
        [
            BillItem(
                bill_id=bill_id,
                product_name=str(item.product_name).strip(),
                quantity=float(item.quantity),
                price=float(item.price),
                total=float(item.quantity * item.price),
            )
            for item in request.items
        ]
    )
    db.commit()

    print(f"RECEIVED PAYLOAD: {payload.json()}")

    stream = generate_bill_excel(
        {
            "bill_id": bill_id,
            "customer_name": customer.name,
            "customer_type": customer.customer_type,
            "phone": customer.phone,
            "location": customer.location,
            "bill_date": datetime.now(),
            "items": [
                {
                    "product_name": item.product_name,
                    "quantity": item.quantity,
                    "price": item.price,
                    "total": item.total,
                }
                for item in payload.items
            ],
            "subtotal": subtotal,
            "packaging_fee": request.packaging_fee,
            "grand_total": grand_total,
            "advance_paid": request.advance_paid,
            "balance_due": balance_due,
        }
    )

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{bill_id}.xlsx"'},
    )


@app.get("/api/bills/{bill_id}/download")
def download_bill(bill_id: str, format: str = "excel", db: Session = Depends(get_db)) -> StreamingResponse:
    bill = db.scalar(
        select(Bill)
        .options(selectinload(Bill.customer), selectinload(Bill.items))
        .where(Bill.bill_id == bill_id)
    )
    if bill is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bill not found.",
        )

    payload = {
        "bill_id": bill.bill_id,
        "customer_name": bill.customer.name if bill.customer else "",
        "customer_type": bill.customer.customer_type if bill.customer else "",
        "phone": bill.customer.phone if bill.customer else None,
        "location": bill.customer.location if bill.customer else None,
        "bill_date": bill.created_at,
        "items": [
            {
                "product_name": item.product_name,
                "quantity": item.quantity,
                "price": item.price,
                "total": item.total,
            }
            for item in bill.items
        ],
        "subtotal": bill.subtotal,
        "packaging_fee": bill.packaging_fee,
        "grand_total": bill.grand_total,
        "advance_paid": bill.advance_paid,
        "balance_due": bill.balance_due,
    }

    if format == "pdf":
        return StreamingResponse(
            generate_pdf_bill(payload),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{bill_id}.pdf"'},
        )

    return StreamingResponse(
        generate_bill_excel(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{bill_id}.xlsx"'},
    )
